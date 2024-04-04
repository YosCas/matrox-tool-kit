import paramiko
from openpyxl import Workbook, load_workbook
import wx
import wx.adv
import wx.lib.filebrowsebutton as filebrowse
import os
import configparser
import time

# Función para establecer una conexión SSH con el servidor remoto
def ssh_connect(ip, username, password):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(ip, username=username, password=password)
    return ssh

# Función para buscar archivos que contienen una palabra clave en un directorio remoto
def search_files_with_keyword(ssh, remote_path, keyword):
    stdin, stdout, stderr = ssh.exec_command(f'ls "{remote_path}"')
    return [f"{remote_path}/{file_name}" for file_name in stdout.read().decode().splitlines() if keyword in file_name]

# Función para leer los nombres de las hojas de un archivo de Excel
def read_sheet_names_from_excel(file_path):
    wb = load_workbook(file_path)
    return [row[0] for row in wb.active.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True) if row[0]]

# Función para obtener el tipo de archivo basado en su extensión
def get_file_type(file_name):
    return {
        '.txt': 'Ver archivo txt',
        '.png': 'Ver imagen png',
        '.jpg': 'Ver imagen jpg'
    }.get(os.path.splitext(file_name)[1], '')

# Función para leer las direcciones IP desde un archivo de Excel
def read_ip_addresses_from_excel(file_path):
    wb = load_workbook(file_path)
    return [row[0] for row in wb.active.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]

# Función para guardar la configuración en un archivo de configuración INI
def save_config(file_path, username, selected_date):
    config = configparser.ConfigParser()
    config['USER_SETTINGS'] = {
        'ConfigFilePath': file_path,
        'Username': username,
        'SelectedDate': selected_date.Format("%Y-%m-%d") if selected_date else ""
    }
    with open('config.ini', 'w') as configfile:
        config.write(configfile)

# Función para crear el archivo de Excel con los archivos encontrados
def create_excel(event):
    config_file_path = file_picker.GetValue()
    username = text_username.GetValue()  # Obtener el nombre de usuario ingresado
    password = text_password.GetValue()  # Obtener la contraseña ingresada
    selected_date = calendar_picker.GetDate()
    remote_path_date = selected_date.Format("%Y-%m-%d")
    keyword = "Fail"
    remote_path = f"/home/mtxuser/Documents/Matrox Design Assistant/SavedImages/{remote_path_date}"
    try:
        ip_addresses = read_ip_addresses_from_excel(config_file_path)
        wb = Workbook()
        for ip, sheet_name in zip(ip_addresses, read_sheet_names_from_excel(config_file_path)):
            try:
                with ssh_connect(ip, username, password) as ssh:
                    files_with_keyword = search_files_with_keyword(ssh, remote_path, keyword)
                    if files_with_keyword:
                        ws = wb.create_sheet(title=sheet_name)
                        ws.append(["File Name", "Link"])
                        for file_path in files_with_keyword:
                            file_name = os.path.basename(file_path)
                            file_type = get_file_type(file_name)
                            link = f"\\\\{ip}\\" + username + '\\Documents' + remote_path.split("/Documents")[1].replace('/', '\\') + '\\' + file_name
                            ws.append([file_name, f'=HYPERLINK("{link}", "{file_type}")'])
                        for col in ws.columns:
                            max_length = max(len(str(cell.value)) for cell in col)
                            ws.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2
                        ws.auto_filter.ref = ws.dimensions
                        ws.auto_filter.add_sort_condition("A2:A1048576")
                    else:
                        print(f"No files containing the keyword found for IP {ip}")
            except Exception as e:
                print(f"An error occurred while processing IP {ip}: {e}")
                continue
        wb.remove(wb.active)
        wb.save("failed_files.xlsx")
        print("Excel file 'failed_files.xlsx' created successfully.")
        save_config(config_file_path, username, selected_date)
        os.startfile("failed_files.xlsx")
    except Exception as e:
        print(f"An error occurred: {e}")

# Crear la aplicación wxPython
app = wx.App(False)
frame = wx.Frame(None, wx.ID_ANY, "Matrox Tool Kit V1", size=(425, 490))
notebook = wx.Notebook(frame)
panel = wx.Panel(notebook, wx.ID_ANY)
sizer = wx.BoxSizer(wx.VERTICAL)
file_picker = filebrowse.FileBrowseButton(panel, labelText="Config File Path:", fileMask="*.xlsx", changeCallback=None, startDirectory=".", toolTip="Select Config File")
sizer.Add(file_picker, 0, wx.ALL | wx.EXPAND, 5)

# Campo de entrada para el nombre de usuario
label_username = wx.StaticText(panel, wx.ID_ANY, "Username:")
sizer.Add(label_username, 0, wx.ALL | wx.EXPAND, 5)
text_username = wx.TextCtrl(panel, wx.ID_ANY, "")
sizer.Add(text_username, 0, wx.ALL | wx.EXPAND, 5)

# Campo de entrada para la contraseña
label_password = wx.StaticText(panel, wx.ID_ANY, "Password:")
sizer.Add(label_password, 0, wx.ALL | wx.EXPAND, 5)
text_password = wx.TextCtrl(panel, wx.ID_ANY, "", style=wx.TE_PASSWORD)
sizer.Add(text_password, 0, wx.ALL | wx.EXPAND, 5)

calendar_picker = wx.adv.CalendarCtrl(panel, wx.ID_ANY)
sizer.Add(calendar_picker, 0, wx.ALL | wx.EXPAND, 5)
text_date = wx.TextCtrl(panel, wx.ID_ANY, "", style=wx.TE_READONLY)
sizer.Add(text_date, 0, wx.ALL | wx.EXPAND, 5)
line = wx.StaticLine(panel, wx.ID_ANY, size=(200, -1), style=wx.LI_HORIZONTAL)
sizer.Add(line, 0, wx.ALL | wx.EXPAND, 5)
btn_create = wx.Button(panel, wx.ID_ANY, "Create Excel")
sizer.Add(btn_create, 0, wx.ALL | wx.ALIGN_CENTER_HORIZONTAL, 5)
btn_create.Bind(wx.EVT_BUTTON, create_excel)
calendar_picker.Bind(wx.adv.EVT_CALENDAR_SEL_CHANGED, lambda event: text_date.SetValue(calendar_picker.GetDate().Format("%Y-%m-%d")))
panel.SetSizer(sizer)
notebook.AddPage(panel, "Img Review")
frame.Show(True)
app.MainLoop()
